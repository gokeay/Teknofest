from openpyxl import load_workbook

# Veri Alınacak ve Aktarılacak Dosyaları Yükle:

# Veri alınacak dosyayı (vald) yükle
vakd_workbook = load_workbook('Z_Raporu.xlsx')  # Dosya adını ve yolunu gerekirse değiştir

# Veri aktarılacak dosyayı (vakd) yükle
vald_workbook = load_workbook('firma1.xlsx')  # Dosya adını ve yolunu gerekirse değiştir

# İşlem yapılacak sayfaları seç
vakd_sheet = vakd_workbook.active
vald_sheet = vald_workbook.active
vakd_sheet["g6"] = "deneme"

# ---------------------------------------------------

# Yer Bilgisine Göre Alanı Belirle:

yer_bilgisi = vald_sheet['F1'].value  # F sütununda bulunan yer bilgisini al

print("yer Bilgisi: ")
print(yer_bilgisi)

if yer_bilgisi == 'VIP':
    baslangic_sutunu = 'A'
    bitis_sutunu = 'Q'
else:  # TSK
    baslangic_sutunu = 'AH'
    bitis_sutunu = 'AW'

# ----------------------------------------------------------------

# İkramlık Bilgisine Göre Alt Alanı Belirle:

ikramli_k_bilgisi = vakd_sheet['G1'].value  # G sütununda bulunan ikramlık bilgisini al
print("ikramlık Bilgisi: ")
print(ikramli_k_bilgisi)

if ikramli_k_bilgisi == 'teslimatGeldi':
    alt_alan = 'F'
elif ikramli_k_bilgisi == 'mevcut':
    alt_alan = 'G'
else:
    alt_alan = 'H' if yer_bilgisi == 'VIP' else 'AM'

# ----------------------------------------------------------------

# Satırı Belirtilen Alana Ekleyin:

# Yeni satırın indeksini belirle
yeni_sira = 5

# Saat bilgisini D sütunundan al ve belirlenen alt alana ekle
vakd_sheet[f'{alt_alan}{yeni_sira}'] = vald_sheet[f'D{yeni_sira}'].value

# Fotoğraf linkini Q sütunundan al ve belirlenen alt alana ekle
vakd_sheet[f'{alt_alan}{yeni_sira+1}'] = vald_sheet[f'Q{yeni_sira}'].value

# Adet bilgisini P sütunundan al ve belirlenen alt alana ekle
vakd_sheet[f'{alt_alan}{yeni_sira+2}'] = vald_sheet[f'P{yeni_sira}'].value

# Değişiklikleri kaydet
vakd_workbook.save('Z_Raporu.xlsx')
